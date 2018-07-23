import pandas as pd
import numpy as np

import datetime
from datetime import date, timedelta

import sys
#sys.path.append('/opt/mnt/publicdrive/Analytics/Gerard/Utils/')
sys.path.append('/Volumes/ugcompanystorage/Company/public/Analytics/Gerard/Utils/')
from GA.GA_obj import GA


from functools import reduce

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border
import os
from itertools import islice
def find_row(sheet):
    row = 1
    col = 1
    active_cell = 0

    while(active_cell != None):
        row = row + 1
        print(row)
        active_cell = (sheet.cell(row=row, column=col)).value
        print (active_cell)

    print('---')
    print(row)
    return row

def ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results, segment, sort):

    ## The GA object takes  a profile ID and the location of your credential file as argument to create the object
    #External = GA('API Key goes here', filelocation='/opt/mnt/publicdrive/Analytics/Gerard/Utils/GA/')
    External = GA('API Key goes here', filelocation='/Volumes/ugcompanystorage/Company/public/Analytics/Gerard/Utils/GA/')
    query_response = External.get_results(start_date=start_date,
                    end_date=end_date,
                    filter_var=(None if filter_var==0 else filter_var),
                    metrics=(None if metrics==0 else metrics),
                    dimensions=(None if dimensions==0 else dimensions),
                    max_results=(None if max_results==0 else max_results),
                    segment=(None if segment==0 else segment),
                    sort=(None if sort==0 else sort))

    # Results are stored in 'rows'
    try:

        query_response['rows']

    # coverts list to a dataframe and grabs the first value from the dataframe
        result = pd.DataFrame(query_response['rows']).iloc[0, 0]

    # result:
        return result

    except:
        return 0
#----------------------------------------------------------------------------------------------------------------------
Todays_Date = pd.to_datetime('today')


Source_Folder_Path = "/Volumes/ugcompanystorage/Company/public/Analytics/Gerard/MarketingProjects/ItemVideoPriorities"
#Source_Folder_Path = "/opt/mnt/publicdrive/Analytics/Gerard/MarketingProjects/ItemVideoPriorities/RevenueAndQuantityData.csv"
os.chdir(Source_Folder_Path)


UGItemData = pd.read_csv("RevenueAndQuantityData.csv")
#Source_Report_XLSX_File = openpyxl.load_workbook("Source_Report_test.xlsx")


#
UGItemDataDataFrame = pd.DataFrame(UGItemData)
#print(SourceReportDataFrameFull)

UGItemDataDataFrameLessColumnHeaders = UGItemDataDataFrame.drop(0,axis=0)



#This counts the number of rows in the dataframe
DataFrame_Number_Rows = int(len(UGItemDataDataFrameLessColumnHeaders.index))
print(DataFrame_Number_Rows)


#This is the master results, we will not give this to the user and it will live in the Analytics Directory pon di W
Results_File = openpyxl.load_workbook('/Users/gconnolly/Desktop/Video_Priorities_Results.xlsx')
Results_Tab = Results_File.get_sheet_by_name('Results')

#trying to work around w drive issues
#Results_File = Workbook ('Video_Priorities_Results')
#worksheet1 = Results_File.create_sheet('Results')
#Results_Tab = worksheet1

for row in Results_Tab['A1:G50000']:
    for cell in row:
        cell.value = None


#column 1 name Item Id
col = 1
Results_Tab.cell(row = 1, column = col).value = str("Item ID")


#column 2 name Item Name
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Item Name")



#column 3 name Live Date
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Live Date")

#column 4 name Total Revenue
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Total Revenue")

#column 5 name Total Unit Sales
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Total Unit Sales")

#column 6 Total Website Page Views
col = col + 1
Results_Tab.cell(row = 1, column = col).value = str("Total Website Page Views")

#column 7 name Average Views Per Month
col = col + 1
Results_Tab.cell(row= 1, column=col).value = str("Average Views Per Month")

#column 8 name Average Revenue Per Month
col = col + 1
Results_Tab.cell(row= 1, column=col).value = str("Average Revenue Per Month")

#column 9 name Average Revenue Per Month
col = col + 1
Results_Tab.cell(row= 1, column=col).value = str("Average Unit Sales Per Month")

#column 10 name Geometric Mean
col = col + 1
Results_Tab.cell(row= 1, column=col).value = str("Geometric Mean of Avg Revenue and Avg Views")

Results_File.save('Video_Priorities_Results.xlsx')



#----------------------------------------------------------------------------------------------------------------------



x=0
while x <= DataFrame_Number_Rows:


        query_parameters = UGItemDataDataFrameLessColumnHeaders.iloc[x]
        print(query_parameters)

        Item_ID = int(query_parameters.iloc[0])

        Total_Sales = query_parameters.iloc[1]
        Total_Quantity_Sold = query_parameters.iloc[2]
        Live_Date_Raw = query_parameters.iloc[3]
        Live_Date = datetime.datetime.strptime(Live_Date_Raw,'%Y-%m-%d')

        Clean_Item_Name_No_Extra_Hyphen = Item_name_raw.split(' -', 1)[0]
        print(Clean_Item_Name_No_Extra_Hyphen)
        # Some Item Names Have Apostrophes, that will not work
        Clean_Item_Name_No_Apostrophe = Clean_Item_Name_No_Extra_Hyphen.replace("'", "")
        print(Clean_Item_Name_No_Apostrophe)
        Clean_Item_Name_No_Comma = Clean_Item_Name_No_Apostrophe.replace(",", "")
        print(Clean_Item_Name_No_Comma)
        Clean_Item_Name_No_Stupid_Periods = Clean_Item_Name_No_Comma.replace("...", "")
        print(Clean_Item_Name_No_Stupid_Periods)
        Clean_Item_Name_No_Period = Clean_Item_Name_No_Stupid_Periods.replace(".", "")
        print(Clean_Item_Name_No_Period)
        Clean_Item_Name_No_Colon = Clean_Item_Name_No_Period.replace(":", "")
        print(Clean_Item_Name_No_Colon)
        Clean_Item_Name_No_Ampersand = Clean_Item_Name_No_Colon.replace("& ", "")
        print(Clean_Item_Name_No_Ampersand)
        Clean_Item_Name_No_Exclaimation_Point = Clean_Item_Name_No_Ampersand.replace("!", "")
        print(Clean_Item_Name_No_Exclaimation_Point)
        Clean_Item_Name_No_BackSlash = Clean_Item_Name_No_Exclaimation_Point.replace("/", " ")
        print(Clean_Item_Name_No_BackSlash)
        Clean_Item_Name_No_Hyphen_Two = Clean_Item_Name_No_BackSlash.replace("- ", " ")
        print(Clean_Item_Name_No_Hyphen_Two)
        Clean_Item_Name_No_Left_Parentheses = Clean_Item_Name_No_Hyphen_Two.replace("(", "")
        print(Clean_Item_Name_No_Left_Parentheses)
        Clean_Item_Name_No_Right_Parentheses = Clean_Item_Name_No_Left_Parentheses.replace(")", "")
        print(Clean_Item_Name_No_Right_Parentheses)
        Clean_Item_Name_No_Question_Mark = Clean_Item_Name_No_Right_Parentheses.replace("?", "")
        print(Clean_Item_Name_No_Question_Mark)
        Clean_Item_Name_No_Plus_Sign = Clean_Item_Name_No_Question_Mark.replace("+", "")
        Clean_Item_Name_No_Plus_Remove_Duplicate_Spaces = Clean_Item_Name_No_Plus_Sign.replace("  ", " ")
        print(Clean_Item_Name_No_Plus_Sign)
        Item_Name_Hyphenated = Clean_Item_Name_No_Plus_Remove_Duplicate_Spaces.replace(" ", "-")
        print(Item_Name_Hyphenated)

        #Calculate Months Between Live Date and Current date
        def diff_month(d1, d2):
            return (d1.year - d2.year) * 12 + d1.month - d2.month

        print(diff_month(Todays_Date,Live_Date))

        Total_Months = (diff_month(Todays_Date,Live_Date))





        Page_Path_To_Site = "ga:pagePath=@/product/{}".format(Item_Name_Hyphenated)

        # metric one Pageviews from blog
        start_date = Live_Date.strftime('%Y-%m-%d')
        end_date = Todays_Date.strftime('%Y-%m-%d')
        filter_var = Page_Path_To_Site
        metrics = 'ga:uniquePageviews'
        dimensions = 0
        max_results = 0
        segment = 0
        sort = 0

        All_Time_Page_Views = ga_query(start_date, end_date, filter_var, metrics, dimensions, max_results, segment,
                                        sort)
        print('\n' + str(x) +'. GA QUERY1 Page_Views_From_Blog: ' + str(All_Time_Page_Views))

        Results_File = openpyxl.load_workbook('Video_Priorities_Results.xlsx')

        Results_Tab = Results_File.get_sheet_by_name('Results')

        empty_row = find_row(Results_Tab)

        # add item URL, column 1 name Item Id
        col = 1
        Results_Tab.cell(row=empty_row, column=col).value = Item_ID


        # column 2 name Item Name
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = Item_name_raw

        # column 3 name Live Date
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = Live_Date.strftime('%Y-%m-%d')

        #column 4 name Total Revenue
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = float(Total_Sales)
        Results_Tab.cell(row=empty_row, column=col).number_format = "$#################"

        # column 5 name Total Unit Sales
        # add number of people that came from the blog and added the item to cart
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = int(Total_Quantity_Sold)
        Results_Tab.cell(row=empty_row, column=col).number_format = "#######"

        # column 6 Total Website Page Views
        # add total transaction revenue from people who came from the blog and added something on the site to cart
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = float(All_Time_Page_Views)
        Results_Tab.cell(row=empty_row, column=col).number_format = "#######"

        # column 7 name Average Views Per Month
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = int(int(All_Time_Page_Views)/int(Total_Months))
        Results_Tab.cell(row=empty_row, column=col).number_format = "#######"

        # column 8 name Average Revenue Per Month
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = int(int(Total_Sales) / int(Total_Months))
        Results_Tab.cell(row=empty_row, column=col).number_format = "#######"

        # column 9 name Average Revenue Per Month
        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = int(int(Total_Quantity_Sold) / int(Total_Months))
        Results_Tab.cell(row=empty_row, column=col).number_format = "#######"

        # column 10 name Geometric Mean

        # ----------------------------------------------------------------------------------------------------------------------
        Average_Rev_Per_Month = int(int(Total_Sales) / int(Total_Months))

        Average_Views_Per_Month = int(All_Time_Page_Views)/int(Total_Months)


        def geometric_mean(nums):
        #     '''
        #         Return the geometric average of nums
        #         @param    list    nums    List of nums to avg
        #         @return   float   Geometric avg of nums
        #     '''
             return (reduce(lambda x, y: x * y, nums)) ** (1.0 / len(nums))
        #
        Geometric_mean = geometric_mean([Average_Rev_Per_Month,Average_Views_Per_Month])
        # ----------------------------------------------------------------------------------------------------------------------



        col = col + 1
        Results_Tab.cell(row=empty_row, column=col).value = int(Geometric_mean)
        Results_Tab.cell(row=empty_row, column=col).number_format = "######"





        Results_File.save('Video_Priorities_Results.xlsx')











        x = x+1
